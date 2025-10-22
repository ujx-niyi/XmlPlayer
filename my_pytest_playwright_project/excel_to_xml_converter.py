import os
import shutil
import sys
import yaml
from datetime import datetime
from pathlib import Path
import json
import openpyxl
from lxml import etree as LT
from playwright.sync_api import sync_playwright
from logger_utils import Logger  # ✅ New import

# -------------------------------
# Utility functions
# -------------------------------
def ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


def load_config(yaml_path: str, section: str, logger):
    """Load configuration section from YAML file."""
    with open(yaml_path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    if section not in cfg:
        logger.error(f"Configuration section '{section}' not found in {yaml_path}")
        raise ValueError(f"Configuration section '{section}' not found in {yaml_path}")
    logger.success(f"Loaded configuration section: {section}")
    return cfg[section]


def build_output_dir(base_dir: str, section: str, logger):
    """Create timestamped output directory."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = os.path.join(base_dir, f"{section}_{timestamp}")
    ensure_dir(out_dir)
    logger.success(f"Created output directory: {out_dir}")
    return out_dir


# -------------------------------
# Transformation functions
# -------------------------------
def excel_to_xml(excel_path: str, xslt_path: str, output_dir: str, logger):
    """Convert Excel workbook into XML using an XSLT transformation."""
    ensure_dir(output_dir)
    logger.section("Excel → XML Transformation")

    print(f"[DEBUG] Checking file: {excel_path}")
    print(f"[DEBUG] Exists? {os.path.exists(excel_path)}", flush=True)


    NS_Y = "urn:schemas-microsoft-com:office:spreadsheet"
    NS_SS = "urn:schemas-microsoft-com:office:spreadsheet"
    LT.register_namespace("y", NS_Y)
    LT.register_namespace("ss", NS_SS)

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    root = LT.Element(f"{{{NS_Y}}}Workbook")

    for sheet_name in wb.sheetnames:
        logger.info(f"Processing sheet: {sheet_name}")
        ws = wb[sheet_name]
        ws_el = LT.SubElement(root, f"{{{NS_Y}}}Worksheet", {f"{{{NS_SS}}}Name": sheet_name})
        table_el = LT.SubElement(ws_el, f"{{{NS_Y}}}Table")

        for row in ws.iter_rows(values_only=True):
            row_el = LT.SubElement(table_el, f"{{{NS_Y}}}Row")
            for cell_value in row:
                cell_el = LT.SubElement(row_el, f"{{{NS_Y}}}Cell")
                data_el = LT.SubElement(cell_el, f"{{{NS_Y}}}Data")
                data_el.text = "" if cell_value is None else str(cell_value)

    temp_xml_path = os.path.join(output_dir, "temp_excel.xml")
    LT.ElementTree(root).write(temp_xml_path, encoding="utf-8", xml_declaration=True, pretty_print=True)
    logger.info(f"Temporary XML created: {temp_xml_path}")

    xslt_tree = LT.parse(xslt_path)
    transform = LT.XSLT(xslt_tree)
    source_tree = LT.parse(temp_xml_path)
    transformed_tree = transform(source_tree)

    final_xml_path = os.path.join(output_dir, "test_cases.xml")
    transformed_tree.write(final_xml_path, encoding="utf-8", xml_declaration=True, pretty_print=True)
    logger.success(f"XML written successfully: {final_xml_path}")

    if os.path.exists(temp_xml_path):
        os.remove(temp_xml_path)
        logger.info("Removed temporary file: temp_excel.xml")

    copied_excel_path = os.path.join(output_dir, os.path.basename(excel_path))
    shutil.copy2(excel_path, copied_excel_path)
    logger.success(f"Copied original Excel workbook: {copied_excel_path}")

    return final_xml_path


def xml_to_html(xml_path: str, xsl_path: str, output_dir: str, logger):
    """Transform XML into HTML using XSL stylesheet."""
    ensure_dir(output_dir)
    html_path = os.path.join(output_dir, "test_cases.html")
    logger.section("XML → HTML Transformation")

    dom = LT.parse(xml_path)
    xslt = LT.parse(xsl_path)
    transform = LT.XSLT(xslt)
    newdom = transform(dom)

    with open(html_path, "wb") as f:
        f.write(LT.tostring(newdom, pretty_print=True, method="html", encoding="utf-8"))

    logger.success(f"HTML report generated: {html_path}")
    return html_path


def html_to_pdf(html_path: str, output_dir: str, logger):
    """Convert an HTML file to PDF using Playwright."""
    pdf_path = os.path.join(output_dir, "test_cases.pdf")
    ensure_dir(output_dir)
    logger.section("HTML → PDF Conversion")

    with sync_playwright() as p:
        browser = p.chromium.launch()
        page = browser.new_page()
        page.goto(f"file:///{os.path.abspath(html_path)}")
        page.pdf(
            path=pdf_path,
            format="A4",
            margin={"top": "1cm", "bottom": "1cm", "left": "1cm", "right": "1cm"},
            print_background=True,
        )
        browser.close()

    logger.success(f"PDF successfully generated: {pdf_path}")
    return pdf_path


# -------------------------------
# Main startup logic
# -------------------------------

def main():
    if len(sys.argv) != 4:
        print("Usage: python excel_to_xml_converter.py <orgid> <project> <role>", flush=True)
        sys.exit(1)

    orgid, project, role = [s.upper() for s in sys.argv[1:4]]
    section_key = f"{orgid}_{project}_{role}"

    # Pre-create a temporary logger until output folder exists
    temp_log_dir = "data/temp_logs"
    ensure_dir(temp_log_dir)
    logger = Logger(temp_log_dir, log_prefix=f"{section_key}_startup")

    logger.section("Initializing Conversion Pipeline")
    yaml_file = "config.yml"
    cfg = load_config(yaml_file, section_key, logger)

    excel_file = cfg.get("excel_source")
    if not excel_file:
        logger.error(f"Missing 'excel_source' entry in config for {section_key}")
        sys.exit(1)

    if not os.path.exists(excel_file):
        logger.error(f"Excel file not found at path: {excel_file}")
        sys.exit(1)

    output_dir = build_output_dir("data/output", section_key, logger)

    # Recreate logger inside final output directory for full run
    logger = Logger(output_dir, log_prefix=section_key)
    logger.section(f"Running conversion for {section_key}")
    logger.info(f"Excel source dir: {excel_file}")
    logger.info(f"Excel input: {excel_file}")
    logger.info(f"Excel→XML stylesheet: {cfg['from_excel_transform_template']}")
    logger.info(f"XML→HTML stylesheet: {cfg['to_html_transform_template']}")
    logger.info(f"Output directory: {output_dir}")

    # Execute the transformation pipeline
    xml_file = excel_to_xml(excel_file, cfg["from_excel_transform_template"], output_dir, logger)
    html_file = xml_to_html(xml_file, cfg["to_html_transform_template"], output_dir, logger)
    html_to_pdf(html_file, output_dir, logger)

    logger.section("All conversions completed successfully")
    logger.info(f"Log file: {logger.path()}")

    latest_pointer = os.path.join("data", "output", f"latest_{section_key}.json")

    try:
        ensure_dir(os.path.dirname(latest_pointer))
        with open(latest_pointer, "w", encoding="utf-8") as f:
            json.dump({"current_output": os.path.basename(output_dir)}, f, indent=2)
        logger.success(f"Updated latest pointer: {latest_pointer} → {output_dir}")
    except Exception as ex:
        logger.error(f"Failed to update latest pointer: {ex}")

    logger.info("Conversion complete — exiting now.")
    sys.exit(0)



if __name__ == "__main__":
    main()

